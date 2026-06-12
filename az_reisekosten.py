"""
AZ & Reisekosten Tab
Arbeitszeiten- und Reisekosten-Verwaltung via Salesforce FSL (Field Service Lightning)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import threading
import datetime
import calendar
import urllib.request
import urllib.parse
import shutil
import webbrowser
from utils import (C, ToolTip, _btn, _ensure,
                   BASE_DIR, RES_DIR, load_json, save_json, sanitize_error)

# openpyxl wird erst beim ersten Excel-Export geladen (lazy import).
# tool.py installiert es bei Bedarf im Hintergrund-Thread beim Start.

# ── Pfade ──────────────────────────────────────────────────────────────────────
# BASE_DIR und RES_DIR kommen aus utils (kein doppeltes os.path-Calculation)
PROFILE_FILE = os.path.join(RES_DIR, "user_profile.json")

# ── Salesforce-Konfiguration (einmalig durch Admin befüllen) ──────────────────
SF_INSTANCE_URL  = "https://koenig-bauer.my.salesforce.com"
KW_DIR         = os.path.join(RES_DIR, "kw_data")
TEMPLATES_DIR  = os.path.join(RES_DIR, "templates")
KUNDEN_FILE    = os.path.join(RES_DIR, "kunden_vorlagen.json")
TMPL_REISE    = os.path.join(TEMPLATES_DIR, "FB_0020_Reisekostenabrechnung.xlsm")
os.makedirs(KW_DIR,        exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)

DAY_NAMES      = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
STATUS_OPTIONS = ["Arbeit", "Krank", "Urlaub", "Kurzarbeit", "GLZ", "Feiertag", "Sonstiges"]
MONTH_NAMES    = ["Januar", "Februar", "März", "April", "Mai", "Juni",
                  "Juli", "August", "September", "Oktober", "November", "Dezember"]
SF_API_VER     = "v59.0"

ALLGEMEIN_CODES = {
    "0010": "Customer Preparation",
    "0020": "Service Hotline",
    "0030": "Maintenance of Rental Equipment",
    "0040": "Department of Rental Equipment",
    "0050": "Cleaning Service",
    "0060": "Internal Meetings / Trainings",
    "0070": "Inventory Service",
    "0080": "Materials Management Activities",
    "0090": "Documentation Activities",
    "0100": "Fixture Construction",
    "0120": "Inspection of Returns",
    "0130": "Proactive Customer Service",
    "0140": "Automotive Workshop / Inspection",
    "0180": "Preparation Training Documents",
    "0190": "Dealer & Subsidiary Support",
    "0200": "Online Training Preparation",
    "2100": "Activities Service Cloud",
}

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────
# load_json / save_json → identisch mit load_json / save_json aus utils

def _current_kw():
    today = datetime.date.today()
    iso   = today.isocalendar()
    return iso[1], iso[0]

def _week_dates(year, week):
    """Gibt 7 date-Objekte für die ISO-KW zurück (Mo–So)."""
    jan4  = datetime.date(year, 1, 4)
    start = jan4 - datetime.timedelta(days=jan4.isoweekday() - 1)
    start += datetime.timedelta(weeks=week - 1)
    return [start + datetime.timedelta(days=i) for i in range(7)]

def _calc_entry_hours(entry: dict) -> str:
    """Netto-Stunden eines einzelnen Eintrags."""
    try:
        s = datetime.datetime.strptime(entry.get("start", ""), "%H:%M")
        e = datetime.datetime.strptime(entry.get("end",   ""), "%H:%M")
        p = int(entry.get("pause", 0) or 0)
        h = max(0.0, (e - s).seconds / 3600 - p / 60)
        return f"{h:.2f}"
    except Exception:
        return ""

def _calc_hours(day: dict) -> str:
    """Tages-Netto-Stunden — summiert Einträge wenn vorhanden, sonst Tag-Ebene.
    Delegiert Einzel-Berechnung an _calc_entry_hours (keine Logik-Duplikation)."""
    entries = [e for e in day.get("entries", [])
               if e.get("start") and e.get("end")]
    if entries:
        total = sum(float(_calc_entry_hours(e) or 0) for e in entries)
        return f"{total:.2f}" if total else ""
    # Fallback: Tag-Ebene hat dieselbe start/end/pause-Struktur wie ein Eintrag
    return _calc_entry_hours(day)

def _build_route(entries, wohnort="Wohnort"):
    """Baut die Tagesroute für FB_0020:
    Wohnort → Standort (Auftr.Nr) → Standort2 (Auftr.Nr2) → Wohnort"""
    if not entries:
        return ""
    stops = []
    start = (entries[0].get("start_punkt") or wohnort).strip()
    stops.append(start)
    for e in entries:
        dienst = e.get("dienst", "Außendienst")
        if dienst in ("Innendienst", "Homeoffice", "Home Office"):
            code = e.get("allg_code", "")
            bez  = ALLGEMEIN_CODES.get(code, "")
            stop = f"{code} – {bez}" if code else "Innendienst"
        else:
            loc   = (e.get("standort") or e.get("kunde_name", "")).strip()
            auftr = e.get("auftr_nr", "").strip()
            stop  = f"{loc} ({auftr})" if loc and auftr else loc or auftr or ""
        if stop and (not stops or stop != stops[-1]):
            stops.append(stop)
    end = (entries[-1].get("end_punkt") or wohnort).strip()
    if end and end != stops[-1]:
        stops.append(end)
    return " → ".join(stops)

def _entry_detail(entry, wohnort="Wohnort"):
    """Detail-String für einen Eintrag (Spalte 'Details' in UI + Excel).
    AD:  Kundenname: Wohnort → Standort (Auftr.Nr) → Wohnort
    ID:  09000 – Dokumentationstätigkeiten
    """
    dienst = entry.get("dienst", "Außendienst")
    if dienst in ("Innendienst", "Homeoffice", "Home Office"):
        code = entry.get("allg_code", "")
        bez  = ALLGEMEIN_CODES.get(code, "")
        return f"{code} – {bez}" if code else "Innendienst"
    else:
        start  = (entry.get("start_punkt") or wohnort).strip()
        loc    = (entry.get("standort") or "").strip()
        kunde  = (entry.get("kunde_name") or "").strip()
        auftr  = entry.get("auftr_nr", "").strip()
        end    = (entry.get("end_punkt") or wohnort).strip()
        loc_or_kunde = loc or kunde
        stop   = f"{loc_or_kunde} ({auftr})" if loc_or_kunde and auftr else loc_or_kunde or auftr or ""
        parts  = [p for p in [start, stop, end] if p]
        route  = " → ".join(parts)
        # Kundenname als Präfix wenn er nicht schon im Stop steckt
        prefix = f"{kunde}: " if kunde and kunde != loc_or_kunde else ""
        return f"{prefix}{route}"

def _str_to_time(s):
    """'08:30' → datetime.time(8, 30) für Excel-Zeitwerte."""
    try:
        h, m = map(int, s.split(':'))
        return datetime.time(h, m)
    except Exception:
        return None

def _get_template(tmpl_path, title):
    """Gibt Vorlagen-Pfad zurück — fragt Nutzer wenn Datei fehlt und kopiert sie."""
    if os.path.isfile(tmpl_path):
        return tmpl_path
    messagebox.showinfo("Vorlage einrichten",
        f"Vorlage '{os.path.basename(tmpl_path)}' nicht gefunden.\n"
        f"Bitte die Original-Excel-Vorlage einmalig auswählen\n"
        f"(wird dann automatisch gespeichert).")
    src = filedialog.askopenfilename(
        title=f"Vorlage wählen: {title}",
        filetypes=[("Excel Makro-Datei", "*.xlsm"), ("Excel", "*.xlsx"), ("Alle", "*.*")])
    if not src:
        return None
    shutil.copy(src, tmpl_path)
    return tmpl_path

# ══════════════════════════════════════════════════════════════════════════════
class AZReisekostenTab(tk.Frame):
    """Haupt-Widget für den AZ & Reisekosten Tab."""

    def __init__(self, parent):
        super().__init__(parent, bg=C["bg"])
        self._sf_token        = None
        self._sf_inst_url     = None
        self._sf_user_id      = None
        self._sf_display_name = ""
        self._sf_ping_id      = None   # after-ID für SF-Ping

        self._profile = load_json(PROFILE_FILE, {})

        kw, yr = _current_kw()
        self._kw_var   = tk.IntVar(value=kw)
        self._year_var = tk.IntVar(value=yr)

        # {iso_date: {status, start, end, pause, entries:[{...}]}}
        self._day_data       = {}
        self._day_rows       = {}
        self._current_dates  = []
        self._selected_date  = None
        self._current_entry_idx = 0

        self._build_ui()
        # SF-Ping läuft immer, idled solange kein Token gesetzt
        self.after(5000, self._sf_ping_loop)

    # ── UI Aufbau ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Kopfzeile
        hdr = tk.Frame(self, bg=C["header"])
        hdr.pack(fill="x")
        tk.Label(hdr, text="  AZ & Reisekosten  —  Salesforce FSL",
                 bg=C["header"], fg=C["accent"],
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=10, pady=7)
        self._sf_dot = tk.Label(hdr, text="●", bg=C["header"], fg=C["red"],
                                font=("Segoe UI", 13))
        self._sf_dot.pack(side="right", padx=(0, 6), pady=7)
        self._sf_lbl = tk.Label(hdr, text="Nicht verbunden",
                                bg=C["header"], fg=C["subtext"],
                                font=("Segoe UI", 8))
        self._sf_lbl.pack(side="right", pady=7)

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True)

        # Linkes Panel (fest 380 px)
        left = tk.Frame(body, bg=C["surface"], width=380)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
        self._build_left(left)

        tk.Frame(body, bg=C["border"], width=1).pack(side="left", fill="y")

        # Rechtes Panel
        right = tk.Frame(body, bg=C["bg"])
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

    # ── Linkes Panel ──────────────────────────────────────────────────────────

    def _build_left(self, parent):
        canvas = tk.Canvas(parent, bg=C["surface"], highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = tk.Frame(canvas, bg=C["surface"])
        cwin  = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_conf(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_conf(e):
            canvas.itemconfig(cwin, width=e.width)
        inner.bind("<Configure>", _on_inner_conf)
        canvas.bind("<Configure>", _on_canvas_conf)
        canvas.bind("<MouseWheel>",
                    lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))

        self._build_profile_section(inner)
        self._build_sf_section(inner)
        self._build_kw_section(inner)
        self._build_export_section(inner)
        tk.Frame(inner, bg=C["surface"], height=16).pack()

    def _section_header(self, parent, text):
        tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", padx=8, pady=(12, 0))
        tk.Label(parent, text=text, bg=C["surface"], fg=C["accent"],
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", padx=12, pady=(4, 2))

    def _field_row(self, parent, label, var, show=None):
        row = tk.Frame(parent, bg=C["surface"])
        row.pack(fill="x", padx=10, pady=2)
        tk.Label(row, text=label, bg=C["surface"], fg=C["subtext"],
                 font=("Segoe UI", 8), width=13, anchor="w").pack(side="left")
        kw = dict(textvariable=var, font=("Segoe UI", 9),
                  bg=C["surface2"], fg=C["text"],
                  insertbackground=C["accent"],
                  relief="flat", bd=0,
                  highlightthickness=1,
                  highlightbackground=C["border"],
                  highlightcolor=C["accent"])
        if show:
            kw["show"] = show
        tk.Entry(row, **kw).pack(side="left", fill="x", expand=True, ipady=4)

    def _build_profile_section(self, parent):
        self._section_header(parent, "👤  Persönliche Daten")
        self._v_vorname  = tk.StringVar(value=self._profile.get("vorname", ""))
        self._v_nachname = tk.StringVar(value=self._profile.get("nachname", ""))
        self._v_wohnort  = tk.StringVar(value=self._profile.get("wohnort", ""))
        self._v_pers_nr  = tk.StringVar(value=self._profile.get("pers_nr", ""))
        self._field_row(parent, "Vorname",     self._v_vorname)
        self._field_row(parent, "Nachname",    self._v_nachname)
        self._field_row(parent, "Wohnort",     self._v_wohnort)
        self._field_row(parent, "Personal-Nr", self._v_pers_nr)
        f = tk.Frame(parent, bg=C["surface"])
        f.pack(fill="x", padx=10, pady=(6, 2))
        save_prof_btn = _btn(f, "💾  Daten speichern",
                             "#2a6644", "#d4f5e4", "#337a50", self._save_profile)
        save_prof_btn.pack(fill="x")
        ToolTip(save_prof_btn,
                 "Speichert Vor-/Nachname, Wohnort, Personal-Nr.\n"
                 "und die Salesforce Session ID dauerhaft.")

    def _build_sf_section(self, parent):
        self._section_header(parent, "☁  Salesforce Verbindung")

        self._v_sf_session = tk.StringVar(value=self._profile.get("sf_session", ""))

        # ── Schritt-für-Schritt-Anleitung ─────────────────────────────────────
        guide = tk.Frame(parent, bg=C["surface2"])
        guide.pack(fill="x", padx=10, pady=(4, 2))
        tk.Label(guide, text="  So erhältst du die Session ID:",
                 bg=C["surface2"], fg=C["accent"],
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(4, 2))
        for step in [
            "1.  VPN verbinden (GlobalProtect)",
            "2.  Button → Salesforce in Chrome öffnen",
            "3.  Mit K&B-Account einloggen (SSO)",
            "4.  F12  →  Application  →  Cookies",
            "     →  koenig-bauer.lightning.force.com",
            "5.  Zeile  sid  anklicken  →  Wert kopieren",
            "6.  Wert unten einfügen  →  Verbinden",
            "Für eine genaue Anleitung → siehe Dokumentation.",
        ]:
            tk.Label(guide, text=step,
                     bg=C["surface2"], fg=C["subtext"],
                     font=("Segoe UI", 9), anchor="w"
                     ).pack(fill="x", padx=8, pady=1)
        tk.Frame(guide, bg=C["surface2"], height=4).pack()

        # ── Browser-Button ────────────────────────────────────────────────────
        bf = tk.Frame(parent, bg=C["surface"])
        bf.pack(fill="x", padx=10, pady=(4, 2))
        open_btn = _btn(bf, "🌐  Salesforce in Chrome öffnen",
                        C["surface2"], C["accent"], C["overlay"],
                        lambda: webbrowser.open(
                            "https://koenig-bauer.lightning.force.com"))
        open_btn.pack(fill="x")
        ToolTip(open_btn,
                 "Öffnet Salesforce im Standard-Browser.\n"
                 "Danach mit K&B-Account einloggen (SSO).")

        # ── Session-ID-Eingabe ────────────────────────────────────────────────
        sid_row = tk.Frame(parent, bg=C["surface"])
        sid_row.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(sid_row, text="Session ID", bg=C["surface"], fg=C["subtext"],
                 font=("Segoe UI", 8), width=13, anchor="w").pack(side="left")
        sid_entry = tk.Entry(sid_row, textvariable=self._v_sf_session,
                             font=("Segoe UI", 9), show="*",
                             bg=C["surface2"], fg=C["text"],
                             insertbackground=C["accent"],
                             relief="flat", bd=0,
                             highlightthickness=1,
                             highlightbackground=C["border"],
                             highlightcolor=C["accent"])
        sid_entry.pack(side="left", fill="x", expand=True, ipady=4)
        ToolTip(sid_entry,
                 "Salesforce Session ID (Cookie 'sid').\n"
                 "F12 → Application → Cookies\n"
                 "→ koenig-bauer.lightning.force.com → sid\n"
                 "Läuft nach ~8 h oder Browser-Neustart ab.")

        # ── Debug: TSE-Felder anzeigen ───────────────────────────────────────
        dbg = tk.Frame(parent, bg=C["surface"])
        dbg.pack(fill="x", padx=10, pady=(4, 0))
        dbg_btn = _btn(dbg, "🔍  TSE-Felder anzeigen (Debug)",
                       C["border"], C["subtext"], C["overlay"],
                       self._sf_show_tse_fields)
        dbg_btn.pack(fill="x")

        # ── Verbinden-Button ──────────────────────────────────────────────────
        f = tk.Frame(parent, bg=C["surface"])
        f.pack(fill="x", padx=10, pady=(6, 2))
        conn_btn = _btn(f, "🔑  Verbinden  (Session ID)",
                        "#1a5280", "#d4eeff", "#1e6090", self._sf_login)
        conn_btn.pack(fill="x")
        ToolTip(conn_btn,
                 "Verbindet mit Salesforce und prüft die Session ID.\n"
                 "Bei Erfolg werden Name und Status grün angezeigt.")

    def _build_kw_section(self, parent):
        self._section_header(parent, "📅  Kalenderwoche")
        kw_row = tk.Frame(parent, bg=C["surface"])
        kw_row.pack(fill="x", padx=10, pady=4)

        def _lbl(t):
            tk.Label(kw_row, text=t, bg=C["surface"], fg=C["subtext"],
                     font=("Segoe UI", 8)).pack(side="left", padx=(4, 2))

        _lbl("KW")
        tk.Spinbox(kw_row, from_=1, to=53, textvariable=self._kw_var,
                   width=4, font=("Segoe UI", 9),
                   bg=C["surface2"], fg=C["text"],
                   buttonbackground=C["overlay"],
                   relief="flat", bd=0,
                   command=self._on_kw_change).pack(side="left")
        _lbl("Jahr")
        tk.Spinbox(kw_row, from_=2020, to=2099, textvariable=self._year_var,
                   width=6, font=("Segoe UI", 9),
                   bg=C["surface2"], fg=C["text"],
                   buttonbackground=C["overlay"],
                   relief="flat", bd=0,
                   command=self._on_kw_change).pack(side="left")

        f = tk.Frame(parent, bg=C["surface"])
        f.pack(fill="x", padx=10, pady=(4, 2))
        sf_btn = _btn(f, "☁  Von Salesforce laden",
                      "#1a5280", "#d4eeff", "#1e6090",
                      self._load_from_sf)
        sf_btn.pack(fill="x", pady=(0, 3))
        ToolTip(sf_btn,
                 "Importiert TimeSheet-Einträge aus Salesforce FSL\n"
                 "für die gewählte KW. Benötigt aktive Verbindung.\n"
                 "Vorhandene Einträge werden ergänzt/überschrieben.")

        kw_save_btn = _btn(f, "💾  KW-Daten speichern",
                           "#2a6644", "#d4f5e4", "#337a50",
                           self._save_kw_data)
        kw_save_btn.pack(fill="x")
        ToolTip(kw_save_btn,
                 "Speichert alle Einträge der aktuellen KW lokal\n"
                 "inkl. Sonstige Kosten. Beim nächsten Öffnen\n"
                 "automatisch wieder geladen.")

    def _build_export_section(self, parent):
        self._section_header(parent, "📤  Ausgabe")
        f = tk.Frame(parent, bg=C["surface"])
        f.pack(fill="x", padx=10, pady=4)
        sz_btn = _btn(f, "📊  Servicezeitenmeldung Excel",
                      "#1a5a58", "#d4f5f3", "#1e6a68",
                      self._generate_stundenblatt)
        sz_btn.pack(fill="x", pady=(0, 3))
        ToolTip(sz_btn,
                 "Erstellt eine Excel-Tabelle mit allen Arbeitstagen\n"
                 "des aktuellen Monats (alle KWs des Monats werden\n"
                 "zusammengeführt). Innendienst wird mitexportiert.")

        rk_btn = _btn(f, "📊  Reisekosten Excel  (FB_0020)",
                      "#1a5a58", "#d4f5f3", "#1e6a68",
                      self._generate_excel_reisekosten)
        rk_btn.pack(fill="x")
        ToolTip(rk_btn,
                 "Füllt die Reisekostenabrechnung (FB_0020) mit den\n"
                 "Einträgen der aktuellen KW. Innendienst wird\n"
                 "übersprungen. Sonstiges → G195 / P195.")

    # ── Rechtes Panel ─────────────────────────────────────────────────────────

    def _build_right(self, parent):
        # Kopf
        rhdr = tk.Frame(parent, bg=C["surface2"])
        rhdr.pack(fill="x")
        self._week_title = tk.Label(rhdr, text="Wochenansicht",
                                    bg=C["surface2"], fg=C["accent"],
                                    font=("Segoe UI", 10, "bold"))
        self._week_title.pack(side="left", padx=12, pady=6)
        self._kw_range_lbl = tk.Label(rhdr, text="",
                                      bg=C["surface2"], fg=C["subtext"],
                                      font=("Segoe UI", 8))
        self._kw_range_lbl.pack(side="left")

        # Spaltenköpfe
        chdr = tk.Frame(parent, bg=C["surface"])
        chdr.pack(fill="x")
        tk.Frame(chdr, bg=C["border"], height=1).pack(fill="x")
        hrow = tk.Frame(chdr, bg=C["surface"])
        hrow.pack(fill="x", padx=4, pady=3)
        for text, w, exp in [
                ("Wt",3,False),("Datum",9,False),("Status",11,False),
                ("Start",7,False),("Ende",7,False),("Pause",7,False),
                ("Std",6,False),("Auftrag",13,False),("Kundenname",22,False),
                ("Details",0,True)]:
            kw = {"anchor": "w", "font": ("Segoe UI", 8, "bold"),
                  "bg": C["surface"], "fg": C["accent"]}
            if w:
                kw["width"] = w
            lbl = tk.Label(hrow, text=text, **kw)
            if exp:
                lbl.pack(side="left", padx=2, fill="x", expand=True)
            else:
                lbl.pack(side="left", padx=2)
        tk.Frame(chdr, bg=C["border"], height=1).pack(fill="x")

        # Wochenzeilen (scrollbar)
        wf     = tk.Frame(parent, bg=C["bg"])
        wf.pack(fill="both", expand=True)
        canvas = tk.Canvas(wf, bg=C["bg"], highlightthickness=0)
        vsb    = ttk.Scrollbar(wf, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        self._week_canvas = canvas
        self._week_inner  = tk.Frame(canvas, bg=C["bg"])
        cwin = canvas.create_window((0, 0), window=self._week_inner, anchor="nw")
        self._week_inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(cwin, width=e.width))
        canvas.bind("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))

        # ── Stunden-Übersicht Bar ─────────────────────────────────────────────
        tk.Frame(parent, bg=C["border"], height=1).pack(fill="x")
        self._stunden_bar = tk.Frame(parent, bg=C["header"])
        self._stunden_bar.pack(fill="x")
        self._stunden_bar_labels = {}
        for key, text, fg in [
            ("month_lbl",  "Monat:",      C["subtext"]),
            ("month_val",  "–",           C["accent"]),
            ("sep1",       "  │  ",       C["border"]),
            ("prev_lbl",   "Vormonat:",   C["subtext"]),
            ("prev_val",   "–",           C["subtext"]),
            ("sep2",       "  │  ",       C["border"]),
            ("gleit_lbl",  "Gleitzeit:",  C["subtext"]),
            ("gleit_val",  "–",           C["yellow"]),
        ]:
            lbl = tk.Label(self._stunden_bar, text=text, bg=C["header"], fg=fg,
                           font=("Segoe UI", 8, "bold" if key.endswith("_val") else "normal"))
            lbl.pack(side="left", padx=(10 if key.endswith("_lbl") else 3, 0), pady=4)
            self._stunden_bar_labels[key] = lbl

        # ── KW-Gesamtkosten (einmalig pro KW, kein Tagesbezug) ───────────────
        self._kw_extras_bar    = tk.Frame(parent, bg=C["surface"])
        self._kw_extras_bar.pack(fill="x")
        self._kw_sonstiges_var     = tk.StringVar()
        self._kw_sonstiges_txt_var = tk.StringVar()

        def _kl(t):
            tk.Label(self._kw_extras_bar, text=t,
                     bg=C["surface"], fg=C["subtext"],
                     font=("Segoe UI", 8)).pack(side="left", padx=(6, 2), pady=3)

        def _ke(var, w):
            tk.Entry(self._kw_extras_bar, textvariable=var, width=w,
                     bg=C["surface2"], fg=C["text"], font=("Segoe UI", 9),
                     relief="flat", bd=0, highlightthickness=1,
                     highlightbackground=C["border"], highlightcolor=C["accent"],
                     insertbackground=C["accent"]
                     ).pack(side="left", padx=(0, 4), ipady=2)

        _kl("Sonstiges €:")
        _ke(self._kw_sonstiges_var, 8)
        _kl("Bezeichnung:")
        _ke(self._kw_sonstiges_txt_var, 22)

        # Detail-Panel (unten) – wird erst beim ersten Tages-Klick gebaut
        tk.Frame(parent, bg=C["border"], height=1).pack(fill="x")
        det = tk.Frame(parent, bg=C["surface"], height=240)
        det.pack(fill="x")
        det.pack_propagate(False)
        self._detail_frame = det
        self._detail_built = False

        self._on_kw_change()

    def _ensure_detail_built(self):
        """Baut das Detail-Panel beim ersten Tages-Klick (lazy)."""
        if self._detail_built:
            return
        self._build_detail_panel(self._detail_frame)
        self._detail_built = True

    def _build_detail_panel(self, parent):
        # Kopfzeile + Vorlagen-Auswahl
        hdr = tk.Frame(parent, bg=C["surface"])
        hdr.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(hdr, text="  Auftrags-Details  (Klick auf Zeile zum Auswählen)",
                 bg=C["surface"], fg=C["accent"],
                 font=("Segoe UI", 8, "bold")).pack(side="left")
        tk.Button(hdr, text="💾 Kunde speichern", bg=C["surface2"], fg=C["green"],
                  activebackground=C["overlay"], activeforeground=C["green"],
                  font=("Segoe UI", 8), relief="flat", bd=0, padx=8,
                  command=self._quick_save_kunde).pack(side="left", padx=(12, 0))

        # Kundenvorlagen
        self._kunden_vorlagen = load_json(KUNDEN_FILE, {})
        self._v_vorlage = tk.StringVar()
        vorl_names = list(self._kunden_vorlagen.keys())
        self._vorl_cb = ttk.Combobox(hdr, textvariable=self._v_vorlage,
                                     values=vorl_names, width=18,
                                     font=("Segoe UI", 8))
        self._vorl_cb.pack(side="right", padx=(4, 0))
        self._vorl_cb.bind("<<ComboboxSelected>>", self._vorlage_laden)
        tk.Label(hdr, text="Vorlage:", bg=C["surface"], fg=C["subtext"],
                 font=("Segoe UI", 8)).pack(side="right", padx=(8, 2))
        tk.Button(hdr, text="＋", bg=C["surface2"], fg=C["accent"],
                  activebackground=C["overlay"], activeforeground=C["accent"],
                  font=("Segoe UI", 9, "bold"), relief="flat", bd=0, padx=6,
                  command=self._vorlage_speichern).pack(side="right", padx=2)

        self._dv_auftrnr   = tk.StringVar()
        self._dv_kunde     = tk.StringVar()
        self._dv_standort  = tk.StringVar()
        self._dv_startpkt  = tk.StringVar(value="Wohnort")
        self._dv_endpkt    = tk.StringVar(value="Wohnort")
        self._dv_allg_code = tk.StringVar(value=next(iter(ALLGEMEIN_CODES)))
        self._dv_start     = tk.StringVar(value="08:00")
        self._dv_end       = tk.StringVar(value="17:00")
        self._dv_pause     = tk.StringVar(value="30")

        def _dlbl(p, t):
            tk.Label(p, text=t, bg=C["surface"], fg=C["subtext"],
                     font=("Segoe UI", 8)).pack(side="left", padx=(6, 2))

        def _dentry(p, var, w=12):
            e = tk.Entry(p, textvariable=var, width=w,
                         bg=C["surface2"], fg=C["text"],
                         font=("Segoe UI", 9), relief="flat", bd=0,
                         highlightthickness=1, highlightbackground=C["border"],
                         highlightcolor=C["accent"],
                         insertbackground=C["accent"])
            e.pack(side="left", padx=(0, 2), ipady=3)
            return e

        def _dcb(p, var, vals, w=8):
            cb = ttk.Combobox(p, textvariable=var, values=vals,
                              width=w, state="readonly", font=("Segoe UI", 9))
            cb.pack(side="left", padx=(0, 6))
            return cb

        # ── Zeitzeile (pro Eintrag) ───────────────────────────────────────────
        time_row = tk.Frame(parent, bg=C["surface"])
        time_row.pack(fill="x", padx=10, pady=2)
        _dlbl(time_row, "Start:");  _dentry(time_row, self._dv_start, 7)
        _dlbl(time_row, "Ende:");   _dentry(time_row, self._dv_end,   7)
        _dlbl(time_row, "Pause (min):"); _dentry(time_row, self._dv_pause, 4)

        # Container für die umschaltbaren Zeilen (row1_ad / row1_id / row2)
        self._rows12 = tk.Frame(parent, bg=C["surface"])
        self._rows12.pack(fill="x")

        # row1_ad: Kunden-Felder (für Außendienst / Ausland)
        self._row1_ad = tk.Frame(self._rows12, bg=C["surface"])
        _dlbl(self._row1_ad, "Auftr.Nr:");   _dentry(self._row1_ad, self._dv_auftrnr,  12)
        _dlbl(self._row1_ad, "Kundenname:"); _dentry(self._row1_ad, self._dv_kunde,    24)
        _dlbl(self._row1_ad, "Standort:");   _dentry(self._row1_ad, self._dv_standort, 18)

        # row1_id: Allgemeinkosten-Code (für Innendienst / HO)
        self._row1_id = tk.Frame(self._rows12, bg=C["surface"])
        _dlbl(self._row1_id, "Allgemeinkosten:")
        _dcb(self._row1_id, self._dv_allg_code, list(ALLGEMEIN_CODES.keys()), 6)
        self._allg_lbl = tk.Label(self._row1_id, text="", bg=C["surface"],
                                  fg=C["subtext"], font=("Segoe UI", 9))
        self._allg_lbl.pack(side="left", padx=(0, 8))
        def _allg_code_changed(*_):
            self._allg_lbl.config(text=ALLGEMEIN_CODES.get(self._dv_allg_code.get(), ""))
        self._dv_allg_code.trace_add("write", _allg_code_changed)
        _allg_code_changed()

        # row2: Start-/Endpunkt (nur AD / Ausland)
        self._row2 = tk.Frame(self._rows12, bg=C["surface"])
        _dlbl(self._row2, "Startpunkt:")
        ttk.Combobox(self._row2, textvariable=self._dv_startpkt,
                     values=["Wohnort", "Hotel"], width=10,
                     state="readonly", font=("Segoe UI", 9)
                     ).pack(side="left", padx=(0, 8))
        _dlbl(self._row2, "→  Endpunkt:")
        ttk.Combobox(self._row2, textvariable=self._dv_endpkt,
                     values=["Wohnort", "Hotel"], width=10,
                     state="readonly", font=("Segoe UI", 9)
                     ).pack(side="left", padx=(0, 12))

        # Initialer Zustand: Außendienst
        self._row1_ad.pack(fill="x", padx=10, pady=2)
        self._row2.pack(fill="x", padx=10, pady=2)

        # ── Reisekosten-Felder (für FB_0020) ─────────────────────────────────
        self._dv_dienst       = tk.StringVar(value="Außendienst")
        self._dv_land         = tk.StringVar(value="DE")
        self._dv_uebernacht   = tk.StringVar(value="nein")
        self._dv_fruehstueck  = tk.StringVar(value="nein")
        self._dv_mittag       = tk.StringVar(value="nein")
        self._dv_abend        = tk.StringVar(value="nein")

        row3 = tk.Frame(parent, bg=C["surface"])
        row3.pack(fill="x", padx=10, pady=2)

        _dlbl(row3, "Art:")
        _dcb(row3, self._dv_dienst, ["Außendienst", "Innendienst", "Ausland"], 11)
        self._dv_dienst.trace_add("write", self._dienst_changed)
        _dlbl(row3, "Land:")
        _dentry(row3, self._dv_land, 4)
        _dlbl(row3, "Übernachtung:")
        _dcb(row3, self._dv_uebernacht, ["nein", "ja"], 5)
        _dlbl(row3, "Früh:")
        _dcb(row3, self._dv_fruehstueck, ["nein", "ja"], 5)
        _dlbl(row3, "Mittag:")
        _dcb(row3, self._dv_mittag, ["nein", "ja"], 5)
        _dlbl(row3, "Abend:")
        _dcb(row3, self._dv_abend, ["nein", "ja"], 5)

        # ── Eintrag-Navigation (mehrere Kunden pro Tag) ───────────────────────
        nav = tk.Frame(parent, bg=C["surface"])
        nav.pack(fill="x", padx=10, pady=(4, 2))

        self._entry_nav_lbl = tk.Label(nav, text="Eintrag  1 / 1",
                                       bg=C["surface"], fg=C["subtext"],
                                       font=("Segoe UI", 8))
        self._entry_nav_lbl.pack(side="left", padx=(0, 6))

        def _nav_btn(text, cmd, fg=C["subtext"], tip=""):
            b = tk.Button(nav, text=text, bg=C["surface2"], fg=fg,
                          activebackground=C["overlay"], activeforeground=C["text"],
                          font=("Segoe UI", 9), relief="flat", bd=0, padx=5,
                          cursor="hand2", command=cmd)
            b.pack(side="left", padx=1)
            if tip:
                ToolTip(b, tip)
            return b

        _nav_btn("◀", self._entry_prev,
                 tip="Vorheriger Eintrag des Tages")
        _nav_btn("▶", self._entry_next,
                 tip="Nächster Eintrag des Tages")
        _nav_btn("＋  Eintrag hinzufügen", self._entry_new, fg=C["green"],
                 tip="Neuen Kunden-/Auftragseintrag für diesen Tag hinzufügen.\n"
                     "Nützlich wenn an einem Tag mehrere Kunden besucht wurden.")
        _nav_btn("🗑", self._entry_delete, fg=C["red"],
                 tip="Aktuellen Eintrag löschen.\n"
                     "Mindestens ein Eintrag bleibt immer erhalten.")

        row4 = tk.Frame(parent, bg=C["surface"])
        row4.pack(fill="x", padx=10, pady=(2, 4))
        ueb_btn = _btn(row4, "✓  Übernehmen",
                       "#2a6644", "#d4f5e4", "#337a50",
                       self._detail_save)
        ueb_btn.pack(side="left")
        ToolTip(ueb_btn,
                 "Übernimmt die eingegebenen Daten in den Tageseintrag.\n"
                 "Nicht vergessen: danach noch KW-Daten speichern!")

    # ── Wochenansicht rendern ─────────────────────────────────────────────────

    def _on_kw_change(self, *_):
        kw   = self._kw_var.get()
        year = self._year_var.get()
        dates = _week_dates(year, kw)
        self._current_dates = dates
        self._week_title.config(text=f"KW {kw} / {year}")
        self._kw_range_lbl.config(
            text=f"  {dates[0].strftime('%d.%m.')} – {dates[-1].strftime('%d.%m.%Y')}")
        kw_file = os.path.join(KW_DIR, f"{year}-W{kw:02d}.json")
        _raw = load_json(kw_file, {})
        self._kw_extras = _raw.pop("_extras", {})
        self._day_data  = _raw
        if hasattr(self, "_kw_sonstiges_var"):
            self._kw_sonstiges_var.set(self._kw_extras.get("sonstiges", ""))
            self._kw_sonstiges_txt_var.set(self._kw_extras.get("sonstiges_txt", ""))
        self._render_week(dates)
        self._update_stunden_bar(year, dates[0].month)

    def _get_month_hours(self, year, month):
        """Summiert alle Netto-Stunden eines Monats aus den KW-JSON-Dateien."""
        total = 0.0
        gleit = 0.0
        SOLL  = 8.0
        cal_days = calendar.monthrange(year, month)[1]
        for dn in range(1, cal_days + 1):
            d  = datetime.date(year, month, dn)
            ds = d.isoformat()
            kw = d.isocalendar()[1]
            kw_file = os.path.join(KW_DIR, f"{year}-W{kw:02d}.json")
            data = load_json(kw_file, {})
            # Aktuelle KW: in-memory Daten haben Vorrang vor Datei
            if kw == self._kw_var.get() and year == self._year_var.get():
                data.update(self._day_data)
            day  = data.get(ds, {})
            if not day:
                continue
            status = day.get("status", "Frei" if d.weekday() >= 5 else "Arbeit")
            h_str  = _calc_hours(day)
            try:
                h = float(h_str) if h_str else 0.0
            except Exception:
                h = 0.0
            total += h
            if d.weekday() < 5:
                if status == "Arbeit":
                    gleit += h - SOLL
                elif status in ("GLZ", "Kurzarbeit"):
                    gleit -= SOLL        # Gleitzeitausgleich / Kurzarbeit zieht 8h ab
                # Urlaub / Krank / Feiertag / Sonstiges: kein Gleitzeiteinfluss
        return total, round(gleit, 2)

    def _update_stunden_bar(self, year, month):
        """Aktualisiert die Stunden-Übersicht Bar mit aktuellem + Vormonat."""
        if not hasattr(self, "_stunden_bar_labels"):
            return
        cur_h, cur_g = self._get_month_hours(year, month)
        # Vormonat
        if month == 1:
            prev_y, prev_m = year - 1, 12
        else:
            prev_y, prev_m = year, month - 1
        prev_h, _ = self._get_month_hours(prev_y, prev_m)

        mn = MONTH_NAMES[month - 1]
        pm = MONTH_NAMES[prev_m - 1]
        L  = self._stunden_bar_labels

        L["month_lbl"].config(text=f"{mn} {year}:")
        L["month_val"].config(text=f"{cur_h:.2f} h")
        L["prev_lbl"].config(text=f"{pm} {prev_y}:")
        L["prev_val"].config(text=f"{prev_h:.2f} h")
        gleit_str = f"+{cur_g:.2f} h" if cur_g > 0 else f"{cur_g:.2f} h"
        gleit_fg  = C["green"] if cur_g > 0 else C["red"] if cur_g < 0 else C["subtext"]
        L["gleit_val"].config(text=gleit_str, fg=gleit_fg)

    def _render_week(self, dates):
        for w in self._week_inner.winfo_children():
            w.destroy()
        self._day_rows = {}

        def _entry_short(e):
            dien = e.get("dienst", "Außendienst")
            if dien in ("Innendienst", "Homeoffice", "Home Office"):
                code = e.get("allg_code", "")
                return ALLGEMEIN_CODES.get(code, code) if code else "Innendienst"
            return e.get("kunde_name", "")

        for i, d in enumerate(dates):
            ds             = d.isoformat()
            day            = self._day_data.get(ds, {})
            is_we          = i >= 5
            bg             = C["surface"] if i % 2 == 0 else C["surface2"]
            day_fg         = C["subtext"] if is_we else C["text"]
            default_status = "Frei" if is_we else "Arbeit"
            entries        = day.get("entries", [])
            wohnort        = self._v_wohnort.get() or "Wohnort"

            # ── Tag-Kopfzeile (Wt, Datum, Status, Kopier-Button) ─────────────
            row = tk.Frame(self._week_inner, bg=bg)
            row.pack(fill="x")

            tk.Label(row, text=DAY_NAMES[i], bg=bg, fg=day_fg,
                     font=("Segoe UI", 9, "bold"), width=3, anchor="w"
                     ).pack(side="left", padx=(8, 2), pady=4)
            tk.Label(row, text=d.strftime("%d.%m.%y"), bg=bg, fg=day_fg,
                     font=("Segoe UI", 9), width=9, anchor="w"
                     ).pack(side="left", padx=2)

            sv = tk.StringVar(value=day.get("status", default_status))
            ttk.Combobox(row, textvariable=sv, values=STATUS_OPTIONS,
                         width=10, state="readonly", font=("Segoe UI", 9)
                         ).pack(side="left", padx=4)
            sv.trace_add("write", lambda *_, d=ds, v=sv: self._day_set(d, "status", v.get()))

            if i < 6:
                next_ds = dates[i + 1].isoformat()
                tk.Button(row, text="↓", bg=bg, fg=C["subtext"],
                          activebackground=C["overlay"], activeforeground=C["text"],
                          font=("Segoe UI", 8), relief="flat", bd=0, padx=4, pady=0,
                          command=lambda s=ds, t=next_ds: self._copy_day(s, t)
                          ).pack(side="right", padx=2)

            # ── Eintrag-Unterzeilen ───────────────────────────────────────────
            show_entries = entries if entries else [{}]   # mindestens leere Zeile
            h_lbl    = None
            all_e_rows = []

            for ei, entry in enumerate(show_entries):
                # e_bg = bg: gleicher Block-Hintergrund wie die Kopfzeile.
                # Kein 3-Stufen-Grau – jeder Tag sieht wie ein einheitlicher Block aus.
                e_bg  = bg
                e_row = tk.Frame(self._week_inner, bg=e_bg)
                all_e_rows.append(e_row)
                e_row.pack(fill="x")

                # Einzug: gleiche Breite wie Wt + Datum + Status-Combobox
                tk.Label(e_row, text="", bg=e_bg, width=3
                         ).pack(side="left", padx=(8, 2))
                tk.Label(e_row, text=f"↳{ei+1}" if len(show_entries) > 1 else "",
                         bg=e_bg, fg=C["subtext"],
                         font=("Segoe UI", 8), width=9, anchor="w"
                         ).pack(side="left", padx=2)
                # Status-Platzhalter (Combobox-Breite ≈ 92px)
                tk.Frame(e_row, bg=e_bg, width=92, height=1).pack(side="left", padx=4)

                # Start / Ende / Pause (inline editierbar)
                start_def = entry.get("start") or day.get("start") or ""
                end_def   = entry.get("end")   or day.get("end")   or ""
                pause_def = entry.get("pause") or day.get("pause") or ""

                def _te(frame, ds_=ds, ei_=ei, key="start", default="", w=7, bg_=C["surface2"]):
                    var = tk.StringVar(value=default)
                    tk.Entry(frame, textvariable=var, width=w,
                             bg=bg_, fg=C["text"], font=("Segoe UI", 9),
                             relief="flat", bd=0,
                             highlightthickness=1, highlightbackground=C["border"],
                             highlightcolor=C["accent"],
                             insertbackground=C["accent"]
                             ).pack(side="left", padx=3, ipady=3)
                    var.trace_add("write",
                        lambda *_, d=ds_, i=ei_, k=key, v=var: self._entry_set(d, i, k, v.get()))
                    return var

                _te(e_row, ds, ei, "start", start_def, 7)
                _te(e_row, ds, ei, "end",   end_def,   7)
                _te(e_row, ds, ei, "pause", pause_def, 5)

                # Stunden dieses Eintrags
                h_text = _calc_entry_hours(entry) if entry.get("start") and entry.get("end") \
                         else _calc_hours(day) if ei == 0 else ""
                h_lbl_e = tk.Label(e_row, text=h_text, bg=e_bg, fg=C["accent"],
                                   font=("Segoe UI", 9, "bold"), width=6, anchor="w")
                h_lbl_e.pack(side="left", padx=3)
                if ei == 0:
                    h_lbl = h_lbl_e   # Haupt-h_lbl für _day_set Aktualisierung

                # Auftrags- / Kunden- / Details-Labels
                _dienst    = entry.get("dienst", "Außendienst")
                _allg_code = entry.get("allg_code", "")
                _day_status = day.get("status", "")

                if _dienst in ("Innendienst", "Homeoffice", "Home Office") and _allg_code:
                    # Overhead-Eintrag: Code in Kundenname, Beschreibung in Details
                    auftr_t   = "—"
                    kunde_t   = _allg_code
                    details_t = ALLGEMEIN_CODES.get(_allg_code, _allg_code)
                    k_fg      = C["accent"]
                else:
                    auftr_t   = entry.get("auftr_nr", "") or "—"
                    kunde_t   = _entry_short(entry) or "—"
                    _detail_raw = _entry_detail(entry, wohnort)
                    if not _detail_raw and _day_status and _day_status != "Arbeit":
                        details_t = _day_status
                    else:
                        details_t = _detail_raw or "—"
                    has_customer = bool(entry.get("kunde_name") or entry.get("auftr_nr"))
                    k_fg = C["accent"] if has_customer else C["subtext"]

                tk.Label(e_row, text=auftr_t,   bg=e_bg, fg=C["subtext"],
                         font=("Segoe UI", 8), width=13, anchor="w"
                         ).pack(side="left", padx=2)
                tk.Label(e_row, text=kunde_t,   bg=e_bg, fg=k_fg,
                         font=("Segoe UI", 8), width=22, anchor="w"
                         ).pack(side="left", padx=2)
                tk.Label(e_row, text=details_t, bg=e_bg, fg=C["subtext"],
                         font=("Segoe UI", 7), anchor="w"
                         ).pack(side="left", padx=2, fill="x", expand=True)

                # Klick → Eintrag in Detail-Panel laden
                for w in e_row.winfo_children():
                    if not isinstance(w, (ttk.Combobox, tk.Entry, tk.Button)):
                        w.bind("<Button-1>", lambda ev, d=ds, idx=ei: self._select_entry(d, idx))
                e_row.bind("<Button-1>", lambda ev, d=ds, idx=ei: self._select_entry(d, idx))

            # Klick auf Kopfzeile → Tag auswählen (erster Eintrag)
            row.bind("<Button-1>", lambda ev, d=ds: self._select_day(d))
            for w in row.winfo_children():
                if not isinstance(w, (ttk.Combobox, tk.Entry, tk.Button)):
                    w.bind("<Button-1>", lambda ev, d=ds: self._select_day(d))

            tk.Frame(self._week_inner, bg=C["border"], height=1).pack(fill="x")

            if h_lbl is None:   # Fallback falls keine Einträge
                h_lbl = tk.Label(self._week_inner, text="", bg=bg)
            self._day_rows[ds] = {"row": row, "e_rows": all_e_rows, "h_lbl": h_lbl, "bg": bg, "sv": sv}

    def _copy_day(self, src_ds, dst_ds):
        """Kopiert alle Daten eines Tages auf den nächsten."""
        src = self._day_data.get(src_ds)
        if not src:
            return
        import copy
        self._day_data[dst_ds] = copy.deepcopy(src)
        self._render_week(self._current_dates)
        self._select_day(dst_ds)

    def _day_set(self, ds, key, value):
        self._day_data.setdefault(ds, {})[key] = value
        info = self._day_rows.get(ds)
        if key in ("start", "end", "pause"):
            if info:
                info["h_lbl"].config(text=_calc_hours(self._day_data[ds]))
        elif key == "status":
            if info:
                _STATUS_LBL = {"Krank": "(Kr)", "Urlaub": "(U)",
                               "Kurzarbeit": "(KA)", "GLZ": "(-8h)",
                               "Feiertag": "(FT)", "Sonstiges": "(So)"}
                if value in _STATUS_LBL:
                    info["h_lbl"].config(text=_STATUS_LBL[value])
                else:
                    info["h_lbl"].config(text=_calc_hours(self._day_data[ds]))
            # Start/Ende/Pause löschen bei Non-Arbeit
            if value != "Arbeit":
                day = self._day_data.get(ds, {})
                day.pop("start", None); day.pop("end", None); day.pop("pause", None)
                for e in day.get("entries", []):
                    e.pop("start", None); e.pop("end", None); e.pop("pause", None)
                self._render_week(self._current_dates)
            if self._current_dates:
                self._update_stunden_bar(
                    self._year_var.get(), self._current_dates[0].month)

    def _entry_set(self, ds, ei, key, value):
        """Setzt einen Wert auf Eintrag-Ebene und synchronisiert Tag-Ebene + Stunden-Label."""
        day     = self._day_data.setdefault(ds, {})
        entries = day.setdefault("entries", [{}])
        if 0 <= ei < len(entries):
            entries[ei][key] = value
        # Tag-Ebene syncen
        if key in ("start", "end", "pause"):
            if entries:
                day["start"] = entries[0].get("start", "")
                day["end"]   = entries[-1].get("end",  "")
                day["pause"] = entries[0].get("pause", "")
            info = self._day_rows.get(ds)
            if info:
                info["h_lbl"].config(text=_calc_hours(day))

    def _recolor_block(self, ds, color):
        """Färbt Header-Zeile + alle Eintrags-Unterzeilen eines Tages ein.
        Entry/Combobox/Button-Widgets werden ausgelassen (eigene Farbe behalten)."""
        info = self._day_rows.get(ds)
        if not info:
            return
        for frame in [info["row"]] + info.get("e_rows", []):
            frame.config(bg=color)
            for child in frame.winfo_children():
                if not isinstance(child, (tk.Entry, ttk.Combobox, tk.Button)):
                    try:
                        child.config(bg=color)
                    except tk.TclError:
                        pass

    def _select_entry(self, ds, ei):
        """Wählt Tag und lade einen bestimmten Eintrag in das Detail-Panel."""
        self._ensure_detail_built()
        self._selected_date     = ds
        self._current_entry_idx = ei
        self._load_entry(ds, ei)
        for d2, info in self._day_rows.items():
            self._recolor_block(d2, C["overlay"] if d2 == ds else info["bg"])

    def _dienst_changed(self, *_):
        """Schaltet Detail-Panel zwischen Kunden-Feldern (AD) und Allgemeinkosten (ID)."""
        is_id = self._dv_dienst.get() in ("Innendienst", "Homeoffice", "Home Office")
        self._row1_ad.pack_forget()
        self._row1_id.pack_forget()
        self._row2.pack_forget()
        if is_id:
            self._row1_id.pack(fill="x", padx=10, pady=2)
        else:
            self._row1_ad.pack(fill="x", padx=10, pady=2)
            self._row2.pack(fill="x", padx=10, pady=2)

    def _select_day(self, ds):
        self._ensure_detail_built()
        # Aktuellen Panel-Inhalt erst speichern, bevor wir den neuen Tag laden
        if self._selected_date and self._selected_date != ds:
            self._detail_save(silent=True)
        self._selected_date     = ds
        self._current_entry_idx = 0
        self._load_entry(ds, 0)
        for d2, info in self._day_rows.items():
            self._recolor_block(d2, C["overlay"] if d2 == ds else info["bg"])

    def _load_entry(self, ds, idx):
        """Lädt Eintrag idx des Tages ds in die Detailfelder."""
        day     = self._day_data.get(ds, {})
        entries = day.get("entries", [{}])
        if not entries:
            entries = [{}]
        idx = max(0, min(idx, len(entries) - 1))
        self._current_entry_idx = idx
        e   = entries[idx]
        wohnort = self._v_wohnort.get() or "Wohnort"
        self._dv_auftrnr    .set(e.get("auftr_nr",    ""))
        self._dv_kunde      .set(e.get("kunde_name",  ""))
        self._dv_standort   .set(e.get("standort",    ""))
        self._dv_startpkt   .set(e.get("start_punkt", wohnort))
        self._dv_endpkt     .set(e.get("end_punkt",   wohnort))
        self._dv_allg_code  .set(e.get("allg_code",   next(iter(ALLGEMEIN_CODES))))
        # Zeiten — Eintrag-Ebene, Fallback auf Tag-Ebene für Altdaten
        self._dv_start      .set(e.get("start", day.get("start", "08:00")))
        self._dv_end        .set(e.get("end",   day.get("end",   "17:00")))
        self._dv_pause      .set(e.get("pause", day.get("pause", "30")))
        # dienst + land sind jetzt pro Eintrag (Fallback: Tag-Ebene für alte Daten)
        self._dv_dienst     .set(e.get("dienst",  day.get("dienst",  "Außendienst")))
        self._dv_land       .set(e.get("land",    day.get("land",    "DE")))
        self._dienst_changed()
        self._dv_uebernacht .set(day.get("uebernacht",   "nein"))
        self._dv_fruehstueck.set(day.get("fruehstueck",  "nein"))
        self._dv_mittag     .set(day.get("mittag",       "nein"))
        self._dv_abend      .set(day.get("abend",        "nein"))
        total = len(entries)
        if hasattr(self, "_entry_nav_lbl"):
            self._entry_nav_lbl.config(
                text=f"Eintrag  {idx + 1} / {total}",
                fg=C["accent"] if total > 1 else C["subtext"])

    def _entry_prev(self):
        ds = self._selected_date
        if ds:
            self._detail_save(silent=True)
            self._load_entry(ds, self._current_entry_idx - 1)

    def _entry_next(self):
        ds = self._selected_date
        if ds:
            self._detail_save(silent=True)
            self._load_entry(ds, self._current_entry_idx + 1)

    def _entry_new(self):
        ds = self._selected_date
        if not ds:
            return
        self._detail_save(silent=True)
        entries = self._day_data.setdefault(ds, {}).setdefault("entries", [{}])
        entries.append({})
        self._load_entry(ds, len(entries) - 1)

    def _entry_delete(self):
        ds = self._selected_date
        if not ds:
            return
        entries = self._day_data.get(ds, {}).get("entries", [])
        if len(entries) <= 1:
            messagebox.showinfo("Hinweis", "Mindestens ein Eintrag muss verbleiben.")
            return
        entries.pop(self._current_entry_idx)
        self._load_entry(ds, max(0, self._current_entry_idx - 1))
        self._render_week(self._current_dates)
        self._select_day(ds)

    def _vorlage_speichern(self):
        """Aktuelle Detail-Felder als benannte Vorlage speichern."""
        name = simpledialog.askstring(
            "Vorlage speichern",
            "Name für diese Vorlage (z.B. 'Kunde Muster GmbH'):",
            parent=self)
        if not name:
            return
        self._kunden_vorlagen[name] = {
            "auftr_nr":   self._dv_auftrnr.get(),
            "kunde_name": self._dv_kunde.get(),
            "standort":   self._dv_standort.get(),
            "allg_code":  self._dv_allg_code.get(),
            "start_punkt":self._dv_startpkt.get(),
            "end_punkt":  self._dv_endpkt.get(),
            "land":       self._dv_land.get(),
            "dienst":     self._dv_dienst.get(),
        }
        save_json(KUNDEN_FILE, self._kunden_vorlagen)
        self._vorl_cb.configure(values=list(self._kunden_vorlagen.keys()))
        self._v_vorlage.set(name)

    def _vorlage_laden(self, _event=None):
        """Gewählte Vorlage in die Detail-Felder laden."""
        name = self._v_vorlage.get()
        v = self._kunden_vorlagen.get(name, {})
        if not v:
            return
        self._dv_auftrnr  .set(v.get("auftr_nr",    ""))
        self._dv_kunde    .set(v.get("kunde_name",   ""))
        self._dv_standort .set(v.get("standort",     ""))
        self._dv_allg_code.set(v.get("allg_code",    next(iter(ALLGEMEIN_CODES))))
        self._dv_startpkt .set(v.get("start_punkt",  "Wohnort"))
        self._dv_endpkt   .set(v.get("end_punkt",    "Wohnort"))
        self._dv_land     .set(v.get("land",          "DE"))
        self._dv_dienst   .set(v.get("dienst",        "Außendienst"))

    def _quick_save_kunde(self):
        """Speichert Kundenname + Standort schnell als Vorlage."""
        name  = self._dv_kunde.get().strip()
        stand = self._dv_standort.get().strip()
        if not name and not stand:
            messagebox.showinfo("Hinweis", "Bitte zuerst Kundenname oder Standort eingeben.")
            return
        default = f"{name} – {stand}" if name and stand else name or stand
        vname = simpledialog.askstring(
            "Kunde speichern",
            "Name für diese Kundenkonfiguration:",
            initialvalue=default, parent=self)
        if not vname:
            return
        self._kunden_vorlagen[vname] = {
            "auftr_nr":   self._dv_auftrnr.get(),
            "kunde_name": name,
            "standort":   stand,
            "allg_code":  self._dv_allg_code.get(),
            "start_punkt":self._dv_startpkt.get(),
            "end_punkt":  self._dv_endpkt.get(),
            "land":       self._dv_land.get(),
            "dienst":     self._dv_dienst.get(),
        }
        save_json(KUNDEN_FILE, self._kunden_vorlagen)
        self._vorl_cb.configure(values=list(self._kunden_vorlagen.keys()))
        self._v_vorlage.set(vname)
        messagebox.showinfo("Gespeichert", f"Kunde gespeichert als: {vname}")

    def _detail_save(self, silent=False):
        ds = self._selected_date
        if not ds:
            if not silent:
                messagebox.showinfo("Hinweis", "Bitte zuerst einen Tag auswählen.")
            return
        wohnort  = self._v_wohnort.get() or "Wohnort"
        day      = self._day_data.setdefault(ds, {})
        entries  = day.setdefault("entries", [{}])
        idx      = max(0, min(self._current_entry_idx, len(entries) - 1))
        # Aktuellen Eintrag updaten
        entries[idx] = {
            "auftr_nr":    self._dv_auftrnr.get(),
            "kunde_name":  self._dv_kunde.get(),
            "standort":    self._dv_standort.get(),
            "start_punkt": self._dv_startpkt.get() or wohnort,
            "end_punkt":   self._dv_endpkt.get()   or wohnort,
            "allg_code":   self._dv_allg_code.get(),
            "start":       self._dv_start.get(),
            "end":         self._dv_end.get(),
            "pause":       self._dv_pause.get(),
            "dienst":      self._dv_dienst.get(),
            "land":        self._dv_land.get() or "DE",
        }
        # Tag-Ebene syncen (erster Eintrag = Tag-Start, letzter = Tag-Ende)
        if entries:
            day["start"] = entries[0].get("start", "")
            day["end"]   = entries[-1].get("end",  "")
            day["pause"] = entries[0].get("pause", "")
        day["uebernacht"]  = self._dv_uebernacht.get()
        day["fruehstueck"] = self._dv_fruehstueck.get()
        day["mittag"]      = self._dv_mittag.get()
        day["abend"]       = self._dv_abend.get()
        if not silent:
            self._render_week(self._current_dates)
            self._select_day(ds)
            if self._current_dates:
                self._update_stunden_bar(
                    self._year_var.get(), self._current_dates[0].month)

    # ── Profil & Daten speichern ──────────────────────────────────────────────

    def _save_profile(self):
        self._profile.update({
            "vorname":     self._v_vorname.get(),
            "nachname":    self._v_nachname.get(),
            "wohnort":     self._v_wohnort.get(),
            "pers_nr":     self._v_pers_nr.get(),
            "sf_session":  self._v_sf_session.get(),
        })
        save_json(PROFILE_FILE, self._profile)
        messagebox.showinfo("Gespeichert", "Profil gespeichert.")

    def _save_kw_data(self):
        kw, year = self._kw_var.get(), self._year_var.get()
        kw_file  = os.path.join(KW_DIR, f"{year}-W{kw:02d}.json")
        save_data = dict(self._day_data)
        save_data["_extras"] = {
            "sonstiges":    self._kw_sonstiges_var.get()     if hasattr(self, "_kw_sonstiges_var")     else "",
            "sonstiges_txt":self._kw_sonstiges_txt_var.get() if hasattr(self, "_kw_sonstiges_txt_var") else "",
        }
        save_json(kw_file, save_data)
        messagebox.showinfo("Gespeichert", f"KW {kw}/{year} gespeichert.")

    # ── Salesforce Auth ───────────────────────────────────────────────────────

    def _sf_login(self):
        self._sf_dot.config(fg=C["yellow"])
        self._sf_lbl.config(text="Verbinde…", fg=C["yellow"])
        self.update_idletasks()
        threading.Thread(target=self._sf_do_login, daemon=True).start()

    def _sf_do_login(self):
        try:
            session_id = self._v_sf_session.get().strip()
            if not session_id:
                raise ValueError("Bitte Session ID eingeben.\n\n"
                                 "Chrome mit VPN → F12 → Application → Cookies\n"
                                 "→ koenig-bauer.lightning.force.com → sid")

            inst = SF_INSTANCE_URL.rstrip("/")

            # Session validieren + User-ID holen via userinfo-Endpoint
            url = f"{inst}/services/oauth2/userinfo"
            req = urllib.request.Request(url)
            req.add_header("Authorization", f"Bearer {session_id}")
            req.add_header("Accept", "application/json")
            with urllib.request.urlopen(req, timeout=15) as resp:
                info = json.loads(resp.read().decode("utf-8"))

            self._sf_token        = session_id
            self._sf_inst_url     = inst
            self._sf_user_id      = info.get("user_id", "")
            self._sf_display_name = info.get("preferred_username",
                                    info.get("name", info.get("email", "Verbunden")))

            self.after(0, lambda: (
                self._sf_dot.config(fg=C["green"]),
                self._sf_lbl.config(text=f"Verbunden  ({self._sf_display_name})",
                                    fg=C["green"])))

        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                self._sf_dot.config(fg=C["red"]),
                self._sf_lbl.config(text="Nicht verbunden", fg=C["subtext"])))
            self.after(0, lambda: messagebox.showerror("Salesforce Login", err))

    # ── Salesforce Auto-Ping (immer aktiv, alle 10s) ─────────────────────────

    def _sf_ping_loop(self):
        """Prüft alle 10s ob die Salesforce-Session noch gültig ist.
        Läuft ab dem ersten Login-Versuch dauerhaft im Hintergrund."""
        if not self._sf_token:
            # Kein Token → idle, nächster Check in 10s
            self._sf_ping_id = self.after(10000, self._sf_ping_loop)
            return

        token = self._sf_token
        inst  = self._sf_inst_url

        def check():
            try:
                url = f"{inst}/services/data/"
                req = urllib.request.Request(url)
                req.add_header("Authorization", f"Bearer {token}")
                req.add_header("Accept", "application/json")
                with urllib.request.urlopen(req, timeout=6) as r:
                    r.read()  # nur Status prüfen, Inhalt ignorieren
                def _ok():
                    self._sf_dot.config(fg=C["green"])
                    self._sf_lbl.config(
                        text=f"Verbunden  ({self._sf_display_name})",
                        fg=C["green"])
                self.after(0, _ok)
            except urllib.error.HTTPError as e:
                if e.code in (401, 403):
                    # Session abgelaufen
                    def _expired():
                        self._sf_token = None
                        self._sf_dot.config(fg=C["red"])
                        self._sf_lbl.config(text="Session abgelaufen", fg=C["red"])
                    self.after(0, _expired)
                # Bei anderen HTTP-Fehlern (5xx etc.) – kein Token löschen,
                # nur Dot auf gelb → vorübergehender Fehler
                else:
                    def _warn():
                        self._sf_dot.config(fg=C["yellow"])
                        self._sf_lbl.config(text="Verbindungsproblem", fg=C["yellow"])
                    self.after(0, _warn)
            except Exception:
                # Netzwerkfehler: gelber Dot, Token bleibt erhalten
                def _net():
                    self._sf_dot.config(fg=C["yellow"])
                    self._sf_lbl.config(text="Verbindungsproblem", fg=C["yellow"])
                self.after(0, _net)

        threading.Thread(target=check, daemon=True).start()
        self._sf_ping_id = self.after(10000, self._sf_ping_loop)

    # ── Salesforce SOQL ───────────────────────────────────────────────────────

    def _sf_describe(self, sobject):
        """Gibt alle Felder eines SObject als {name: type} zurück."""
        if not self._sf_token:
            raise RuntimeError("Nicht angemeldet.")
        url = f"{self._sf_inst_url}/services/data/{SF_API_VER}/sobjects/{sobject}/describe"
        req = urllib.request.Request(url)
        req.add_header("Authorization", f"Bearer {self._sf_token}")
        req.add_header("Accept", "application/json")
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return {f["name"]: f["type"] for f in data.get("fields", [])}

    def _sf_show_tse_fields(self):
        """Debug: zeigt Type-Picklist-Werte + Zeitzone-Info von TimeSheetEntry."""
        if not self._sf_token:
            messagebox.showwarning("Hinweis", "Bitte zuerst verbinden.")
            return
        def _run():
            try:
                # Picklist-Werte des Type-Felds aus describe
                url = (f"{self._sf_inst_url}/services/data/{SF_API_VER}"
                       f"/sobjects/TimeSheetEntry/describe")
                req = urllib.request.Request(url)
                req.add_header("Authorization", f"Bearer {self._sf_token}")
                req.add_header("Accept", "application/json")
                with urllib.request.urlopen(req, timeout=30) as resp:
                    desc = json.loads(resp.read().decode("utf-8"))
                type_field = next((f for f in desc.get("fields", [])
                                   if f["name"] == "Type"), None)
                pv = ([v["value"] for v in type_field.get("picklistValues", [])]
                      if type_field else [])

                # Echte Werte aus den letzten 5 Einträgen
                recs = self._sf_query(
                    "SELECT StartTime,Type FROM TimeSheetEntry "
                    "ORDER BY StartTime DESC LIMIT 10")
                sample = "\n".join(
                    f"  {r.get('StartTime','')[:16]}  →  Type='{r.get('Type','')}'"
                    for r in recs)

                # WorkOrder-Felder für Standort suchen
                wo_fields = self._sf_describe("WorkOrder")
                wo_loc = {n: t for n, t in wo_fields.items()
                          if any(k in n.lower() for k in
                                 ("location", "standort", "site", "address",
                                  "city", "street", "ort"))}

                msg = (f"Type Picklist-Werte (API):\n  {pv}\n\n"
                       f"Letzte 10 Einträge:\n{sample}\n\n"
                       f"WorkOrder Standort-Felder:\n"
                       + "\n".join(f"  {n}: {t}" for n, t in sorted(wo_loc.items())))
                self.after(0, lambda: messagebox.showinfo("TSE Type-Werte", msg))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Fehler", str(e)))
        threading.Thread(target=_run, daemon=True).start()

    def _sf_query(self, soql):
        if not self._sf_token:
            raise RuntimeError("Nicht angemeldet.")
        url = (f"{self._sf_inst_url}/services/data/{SF_API_VER}"
               f"/query?q={urllib.parse.quote(soql)}")
        req = urllib.request.Request(url)
        req.add_header("Authorization", f"Bearer {self._sf_token}")
        req.add_header("Accept", "application/json")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8")).get("records", [])
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            try:
                err_data = json.loads(body)
                if isinstance(err_data, list) and err_data:
                    sf_msg = (f"{err_data[0].get('errorCode','')}: "
                              f"{err_data[0].get('message','')}")
                else:
                    sf_msg = str(err_data)[:300]
            except Exception:
                sf_msg = body[:300]
            raise RuntimeError(
                f"HTTP {e.code} von Salesforce:\n{sf_msg}\n\nQuery: {soql[:200]}"
            ) from None

    def _load_from_sf(self):
        if not self._sf_token:
            messagebox.showwarning("Hinweis", "Bitte zuerst bei Salesforce anmelden.")
            return
        dates     = _week_dates(self._year_var.get(), self._kw_var.get())
        start_str = dates[0].isoformat()
        end_str   = dates[-1].isoformat()
        threading.Thread(target=self._sf_do_load,
                         args=(start_str, end_str), daemon=True).start()

    def _sf_do_load(self, start_str, end_str):
        try:
            # 1. TimeSheets
            ts_recs = self._sf_query(
                f"SELECT Id FROM TimeSheet "
                f"WHERE OwnerId='{self._sf_user_id}' "
                f"AND StartDate>={start_str} AND EndDate<={end_str}")
            if not ts_recs:
                self.after(0, lambda: messagebox.showinfo(
                    "SF Laden", "Keine TimeSheets für diese Woche gefunden."))
                return

            ts_ids = "','".join(r["Id"] for r in ts_recs)

            # 2. TimeSheetEntries mit Kundendaten + Overhead-Code
            te_recs = self._sf_query(
                f"SELECT Id,StartTime,EndTime,Type,"
                f"WorkOrderId,formulaAccount__c,OverheadsType__c "
                f"FROM TimeSheetEntry "
                f"WHERE TimeSheetId IN ('{ts_ids}')")

            # 3. WorkOrders (Auftragsnummer + Standort)
            wo_ids = list({r["WorkOrderId"] for r in te_recs
                           if r.get("WorkOrderId")})
            wo_map = {}
            if wo_ids:
                wo_str  = "','".join(wo_ids)
                wo_recs = self._sf_query(
                    f"SELECT Id,WorkOrderNumber,Street,City "
                    f"FROM WorkOrder WHERE Id IN ('{wo_str}')")
                wo_map = {r["Id"]: r for r in wo_recs}

            # ── Zeitstempel parsen ────────────────────────────────────────────
            wohnort   = self._v_wohnort.get() or "Wohnort"
            te_parsed = []
            for te in te_recs:
                raw_s = te.get("StartTime", "")
                if not raw_s:
                    continue
                try:
                    dt_s = datetime.datetime.fromisoformat(raw_s[:19])
                    # Salesforce speichert UTC → lokale Zeit (UTC+2 CEST / UTC+1 CET)
                    import time as _time
                    utc_offset = datetime.timedelta(seconds=-_time.timezone)
                    if _time.daylight and _time.localtime().tm_isdst:
                        utc_offset = datetime.timedelta(seconds=-_time.altzone)
                    dt_s += utc_offset
                except Exception:
                    continue
                try:
                    dt_e = datetime.datetime.fromisoformat(te["EndTime"][:19])
                    dt_e += utc_offset
                except Exception:
                    dt_e = dt_s
                wo_id  = te.get("WorkOrderId") or ""
                wo     = wo_map.get(wo_id, {})
                city   = wo.get("City") or ""
                street = wo.get("Street") or ""
                standort = city if city else street
                te_parsed.append({
                    "date":          dt_s.date().isoformat(),
                    "dt_s":          dt_s,
                    "dt_e":          dt_e,
                    "type":          te.get("Type", "Work"),
                    "wo_id":         wo_id,
                    "auftr_nr":      wo.get("WorkOrderNumber", ""),
                    "kunde_name":    te.get("formulaAccount__c") or "",
                    "standort":      standort,
                    "overheads_type": te.get("OverheadsType__c") or "",
                })
            te_parsed.sort(key=lambda x: (x["date"], x["dt_s"]))

            # ── Pro Tag gruppieren ────────────────────────────────────────────
            from collections import defaultdict
            per_day = defaultdict(list)
            for t in te_parsed:
                per_day[t["date"]].append(t)

            new_data = {}
            for date_s, tes in sorted(per_day.items()):
                # Break-Einträge → Pausenminuten aus Start/Ende berechnen
                pause_tes = [t for t in tes if t["type"] == "Break"]
                work_tes  = [t for t in tes if t["type"] != "Break"]
                pause_min = sum(
                    max(0, int((t["dt_e"] - t["dt_s"]).seconds / 60))
                    for t in pause_tes
                )

                # Tag-Spanne: alle Einträge (inkl. Reisezeit)
                day_start = min(t["dt_s"] for t in tes)
                day_end   = max(t["dt_e"] for t in tes)

                # Arbeits-/Reisezeit nach WorkOrderId gruppieren
                groups   = defaultdict(list)
                anon_idx = 0
                for t in work_tes:
                    key = t["wo_id"] if t["wo_id"] else f"_anon_{anon_idx}"
                    if not t["wo_id"]:
                        anon_idx += 1
                    groups[key].append(t)

                entries_list = []
                for wo_key, group in groups.items():
                    g_start = min(t["dt_s"] for t in group)
                    g_end   = max(t["dt_e"] for t in group)
                    rep     = group[0]
                    overhead = rep.get("overheads_type", "")
                    entries_list.append((g_start, {
                        "start":       g_start.strftime("%H:%M"),
                        "end":         g_end.strftime("%H:%M"),
                        "pause":       "0",   # wird unten für ersten Eintrag gesetzt
                        "start_punkt": wohnort,
                        "end_punkt":   wohnort,
                        "dienst":      "Außendienst" if rep.get("wo_id") else "Innendienst",
                        "auftr_nr":    rep.get("auftr_nr",   ""),
                        "kunde_name":  rep.get("kunde_name", ""),
                        "standort":    rep.get("standort",   ""),
                        "allg_code":   overhead,
                    }))

                entries_list.sort(key=lambda x: x[0])
                entries = [e for _, e in entries_list]
                # Pause nur beim ersten (frühesten) Eintrag des Tages
                if entries:
                    entries[0]["pause"] = str(pause_min)

                # Nur laden wenn mindestens ein echter Kundeneinsatz (WorkOrder) vorhanden
                if not any(t["wo_id"] for t in work_tes):
                    continue

                new_data[date_s] = {
                    "status":  "Arbeit",
                    "start":   day_start.strftime("%H:%M"),
                    "end":     day_end.strftime("%H:%M"),
                    "pause":   str(pause_min),
                    "entries": entries,
                }

            self._day_data.update(new_data)
            self.after(0, lambda: self._render_week(self._current_dates))
            self.after(0, lambda: messagebox.showinfo(
                "SF Laden",
                f"Geladen: {len(new_data)} Tage.\n\n"
                f"⚠  Nicht vergessen: KW-Daten speichern!"))

        except Exception as e:
            err = str(e)
            self.after(0, lambda: messagebox.showerror("SF Ladefehler", err))

    # ── Servicezeitenmeldung Excel ────────────────────────────────────────────

    def _generate_stundenblatt(self):
        try:
            import openpyxl  # noqa: F401  lazy – vermeidet 133 ms beim App-Start
        except ImportError:
            _ensure(("openpyxl", "openpyxl"))
            try:
                import openpyxl  # noqa: F401
            except ImportError:
                messagebox.showerror("Fehler", "openpyxl nicht installiert.\npip install openpyxl")
                return
        kw, year = self._kw_var.get(), self._year_var.get()
        dates    = _week_dates(year, kw)
        month    = dates[0].month
        mn       = MONTH_NAMES[month - 1]
        name_str = f"{self._v_nachname.get()}_{self._v_vorname.get()}"
        dest = filedialog.asksaveasfilename(
            title="Servicezeitenmeldung speichern",
            initialfile=f"Servicezeitenmeldung_{mn}_{year}_{name_str}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Alle", "*.*")])
        if not dest:
            return
        # Snapshot des aktuellen UI-State (thread-sicher kopieren)
        current_day_data = dict(self._day_data)
        # File-I/O komplett im Hintergrundthread – kein GUI-Freeze
        threading.Thread(
            target=self._do_stundenblatt,
            args=(dest, current_day_data, month, year, mn),
            daemon=True).start()

    def _do_stundenblatt(self, dest, current_day_data, month, year, month_name):
        try:
            # ── KW-Daten des Monats hier im Thread laden (kein GUI-Freeze) ────
            month_data = {}
            for w in range(1, 54):
                d = load_json(os.path.join(KW_DIR, f"{year}-W{w:02d}.json"), {})
                for ds, day in d.items():
                    try:
                        if (datetime.date.fromisoformat(ds).month == month and
                                datetime.date.fromisoformat(ds).year == year):
                            month_data[ds] = day
                    except Exception:
                        pass
            # UI-Snapshot hat Vorrang (aktuelle Woche nicht gespeichert)
            month_data.update({
                ds: day for ds, day in current_day_data.items()
                if datetime.date.fromisoformat(ds).month == month
                and datetime.date.fromisoformat(ds).year == year
            })

            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = f"{month_name} {year}"

            # ── Farben ────────────────────────────────────────────────────────
            CLR = {
                "header_bg":   "1E3A5F",   # Dunkelblau Header
                "header_fg":   "FFFFFF",
                "sub_bg":      "2D5986",   # Hellblau Unterzeile
                "sub_fg":      "FFFFFF",
                "ad":          "D6E4F0",   # Außendienst  — blau
                "ad_dark":     "AED6F1",
                "id":          "E8F5E9",   # Innendienst  — grün
                "id_dark":     "C8E6C9",
                "krank":       "FDECEA",   # Krank         — rot
                "urlaub":      "FFF9C4",   # Urlaub        — gelb
                "frei":        "F5F5F5",   # Wochenende    — grau
                "we":          "EEEEEE",
                "total_bg":    "1E3A5F",
                "total_fg":    "FFFFFF",
                "border":      "BDBDBD",
                "accent":      "2196F3",
            }

            def fill(hex_color):
                return PatternFill("solid", fgColor=hex_color)

            def font(bold=False, color="000000", size=10, italic=False):
                return Font(name="Calibri", bold=bold, color=color,
                            size=size, italic=italic)

            def align(h="left", v="center", wrap=False):
                return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

            thin = Side(style="thin", color=CLR["border"])
            thick = Side(style="medium", color="1E3A5F")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
            border_thick_bottom = Border(left=thin, right=thin, top=thin,
                                         bottom=thick)

            # ── Spaltenbreiten ────────────────────────────────────────────────
            col_widths = {
                "A": 6,    # Nr
                "B": 12,   # Datum
                "C": 5,    # Wt
                "D": 16,   # Dienstart
                "E": 10,   # Start
                "F": 10,   # Ende
                "G": 8,    # Pause
                "H": 10,   # Netto Std
                "I": 14,   # Auftr.Nr
                "J": 24,   # Kundenname
                "K": 10,   # Kunden-Nr
                "L": 20,   # Standort
                "M": 30,   # Route
            }
            for col, w in col_widths.items():
                ws.column_dimensions[col].width = w

            # ── Logo-Bereich / Titel-Header (Zeilen 1–5) ──────────────────────
            ws.row_dimensions[1].height = 10
            ws.row_dimensions[2].height = 32
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 18
            ws.row_dimensions[5].height = 10

            # Zeile 2: Haupttitel
            ws.merge_cells("A2:H2")
            c = ws["A2"]
            c.value = "Servicezeitenmeldung"
            c.font  = Font(name="Calibri", bold=True, size=18, color=CLR["header_fg"])
            c.fill  = fill(CLR["header_bg"])
            c.alignment = align("left", "center")
            for col in "BCDEFGH":
                ws[f"{col}2"].fill = fill(CLR["header_bg"])

            ws.merge_cells("I2:M2")
            c2 = ws["I2"]
            c2.value = "K&B Coding GmbH — Servicetechniker"
            c2.font  = Font(name="Calibri", bold=False, size=11,
                            color="CCDDEE", italic=True)
            c2.fill  = fill(CLR["header_bg"])
            c2.alignment = align("right", "center")
            for col in "JKLM":
                ws[f"{col}2"].fill = fill(CLR["header_bg"])

            # Zeile 3: Personaldaten
            name    = f"{self._v_vorname.get()} {self._v_nachname.get()}".strip()
            pers_nr = self._v_pers_nr.get()

            ws.merge_cells("A3:D3")
            ws["A3"].value     = f"  Name:  {name}"
            ws["A3"].font      = font(bold=True, size=11)
            ws["A3"].fill      = fill(CLR["sub_bg"])
            ws["A3"].font      = Font(name="Calibri", bold=True, size=11,
                                      color=CLR["header_fg"])
            ws["A3"].alignment = align("left", "center")
            for col in "BCD":
                ws[f"{col}3"].fill = fill(CLR["sub_bg"])

            ws.merge_cells("E3:H3")
            ws["E3"].value     = f"Personal-Nr:  {pers_nr}"
            ws["E3"].font      = Font(name="Calibri", bold=True, size=11,
                                      color=CLR["header_fg"])
            ws["E3"].fill      = fill(CLR["sub_bg"])
            ws["E3"].alignment = align("center", "center")
            for col in "FGH":
                ws[f"{col}3"].fill = fill(CLR["sub_bg"])

            ws.merge_cells("I3:O3")
            ws["I3"].value     = f"Monat:  {month_name} {year}"
            ws["I3"].font      = Font(name="Calibri", bold=True, size=11,
                                      color=CLR["header_fg"])
            ws["I3"].fill      = fill(CLR["sub_bg"])
            ws["I3"].alignment = align("right", "center")
            for col in ["J","K","L","M","N","O"]:
                ws[f"{col}3"].fill = fill(CLR["sub_bg"])

            # Zeile 4: Legende
            ws.merge_cells("A4:O4")
            ws["A4"].value = (
                "  🔵 Außendienst (AD)     🟢 Innendienst / Home Office (ID/HO)"
                "     🔴 Krank     🟡 Urlaub / Sonderurlaub     ⬜ Wochenende / Frei")
            ws["A4"].font      = Font(name="Calibri", size=9,
                                      color="555555", italic=True)
            ws["A4"].alignment = align("left", "center")

            # ── Spaltenköpfe (Zeile 6) ────────────────────────────────────────
            # Spalten: A-J = Tagesdaten, K = Soll, L = Gleitzeit,
            #          M = Eintrag#, N-R = Kundendaten
            col_widths2 = {
                "A": 5, "B": 12, "C": 5, "D": 18, "E": 8, "F": 8,
                "G": 7, "H": 9, "I": 9, "J": 10,
                "K": 5, "L": 14, "M": 10, "N": 10, "O": 22,
            }
            for col, w in col_widths2.items():
                ws.column_dimensions[col].width = w

            ws.row_dimensions[6].height = 30
            headers = [
                "Nr", "Datum", "Wt", "Dienstart", "Start", "Ende",
                "Pause\n(min)", "Netto\n(h)", "Soll\n(h)", "Gleitzeit\n(+/−)",
                "Eintrag\n#", "Auftrag-Nr", "Kundenname", "Standort",
                "Details",
            ]
            for ci, h in enumerate(headers, start=1):
                cell = ws.cell(row=6, column=ci, value=h)
                cell.font      = Font(name="Calibri", bold=True, size=10,
                                      color=CLR["header_fg"])
                cell.fill      = fill(CLR["header_bg"])
                cell.alignment = align("center", "center", wrap=True)
                cell.border    = Border(
                    left=Side(style="medium", color="FFFFFF"),
                    right=Side(style="medium", color="FFFFFF"),
                    bottom=Side(style="medium", color="5B8DB8"))

            # Gleitzeit-Kopf in Gold hervorheben
            ws.cell(6, 10).fill = fill("7B5EA7")

            ws.auto_filter.ref = "A6:O6"
            ws.freeze_panes    = "A7"

            # ── Daten (Zeilen 7+) ─────────────────────────────────────────────
            SOLL_H   = 8.0   # Standard-Sollzeit
            cal_days = calendar.monthrange(year, month)[1]
            data_row = 7
            ad_days  = 0
            id_days  = 0
            total_h  = 0.0
            gleit_h  = 0.0   # Gleitzeitkonto

            for dn in range(1, cal_days + 1):
                d     = datetime.date(year, month, dn)
                ds    = d.isoformat()
                day   = month_data.get(ds, {})
                wd    = DAY_NAMES[d.weekday()]
                is_we = d.weekday() >= 5

                status  = day.get("status", "Frei" if is_we else "Arbeit")
                dienst  = day.get("dienst", "Außendienst")
                start   = day.get("start", "")
                end     = day.get("end",   "")
                pause   = day.get("pause", "")
                h_str   = _calc_hours(day)
                entries = day.get("entries") or [{}]

                # Netto-Stunden
                try:
                    netto = float(h_str) if h_str else 0.0
                except Exception:
                    netto = 0.0

                # Sollzeit und Gleitzeit berechnen
                is_arbeits = status in ("Arbeit",) and not is_we
                soll  = SOLL_H if is_arbeits else 0.0
                gleit = round(netto - soll, 2) if is_arbeits else 0.0

                # Tagesstatus bestimmt Basis-Farbe
                if status == "Krank":
                    day_art  = "Krank"
                    day_fill = fill(CLR["krank"])
                    soll = SOLL_H; gleit = 0.0
                elif status in ("Urlaub", "Feiertag", "Sonstiges"):
                    day_art  = status
                    day_fill = fill(CLR["urlaub"])
                    soll = SOLL_H; gleit = 0.0
                elif status in ("GLZ", "Kurzarbeit"):
                    day_art  = f"{status} (−8h)"
                    day_fill = fill(CLR["urlaub"])
                    soll = SOLL_H; gleit = -SOLL_H
                elif is_we or status in ("Frei", "GLZ"):
                    day_art  = ""
                    day_fill = fill(CLR["we"])
                else:
                    day_art  = None   # kommt pro Eintrag
                    day_fill = None

                total_h += netto
                gleit_h  = round(gleit_h + gleit, 2)

                gleit_str = (f"+{gleit:.2f}" if gleit > 0
                             else f"{gleit:.2f}" if gleit < 0 else "0")
                gleit_color = ("006100" if gleit > 0
                               else "9C0006" if gleit < 0 else "555555")

                def _entry_fill(e, ei):
                    """Farbe + Art-Text je nach Eintrag-Dienstart."""
                    if day_fill is not None:
                        return day_art, day_fill
                    ed = e.get("dienst", "Außendienst")
                    if ed in ("Innendienst", "Homeoffice", "Home Office"):
                        return "Innendienst / HO", fill(CLR["id"] if ei%2==0 else CLR["id_dark"])
                    elif ed == "Ausland":
                        return "Ausland (AD)", fill("D5E8D4")
                    else:
                        return "Außendienst", fill(CLR["ad"] if ei%2==0 else CLR["ad_dark"])

                # Statistik: AD/ID-Tage aus allen Einträgen zählen
                for e in entries:
                    ed = e.get("dienst", "Außendienst")
                    if day_fill is None:   # nur echte Arbeitstage
                        if ed in ("Innendienst", "Homeoffice", "Home Office"):
                            id_days += 1
                        else:
                            ad_days += 1
                if len(entries) > 1:
                    ad_days -= (len(entries) - 1)   # Tag zählt nur einmal

                # ── Zeilen für jeden Eintrag ──────────────────────────────────
                for ei, entry in enumerate(entries):
                    ws.row_dimensions[data_row].height = 18
                    art_text, row_fill = _entry_fill(entry, ei)

                    # Details: pro Eintrag individueller Detail-String
                    wohnort_val = self._v_wohnort.get() or "Wohnort"
                    route = _entry_detail(entry, wohnort_val)

                    # Eintrag-Zeiten (per-Entry mit Fallback auf Tag-Ebene für ersten Eintrag)
                    e_start = entry.get("start") or (start if ei == 0 else "")
                    e_end   = entry.get("end")   or (end   if ei == 0 else "")
                    e_pause = entry.get("pause") or (pause if ei == 0 else "")
                    e_h_str = _calc_entry_hours(entry)
                    try:
                        e_netto = float(e_h_str) if e_h_str else (netto if ei == 0 else 0.0)
                    except Exception:
                        e_netto = netto if ei == 0 else 0.0

                    if ei == 0:
                        day_vals = [dn, d.strftime("%d.%m.%Y"), wd, art_text,
                                    e_start, e_end,
                                    int(e_pause) if e_pause and str(e_pause).isdigit() else "",
                                    f"{e_netto:.2f}" if e_netto else "",
                                    f"{soll:.1f}" if soll else "",
                                    gleit_str]
                    else:
                        day_vals = ["", d.strftime("%d.%m.%Y"), "", art_text,
                                    e_start, e_end,
                                    int(e_pause) if e_pause and str(e_pause).isdigit() else "",
                                    f"{e_netto:.2f}" if e_netto else "",
                                    "", ""]

                    row_vals = day_vals + [
                        ei + 1,
                        entry.get("auftr_nr",   ""),
                        entry.get("kunde_name", ""),
                        entry.get("standort",   ""),
                        route,
                    ]

                    for ci, val in enumerate(row_vals, start=1):
                        cell = ws.cell(row=data_row, column=ci, value=val)
                        cell.fill   = row_fill
                        cell.border = border_all
                        cell.font   = Font(
                            name="Calibri", size=10,
                            bold=(ci == 4),
                            color="333333" if not is_we else "888888")
                        wrap = (ci == 15)   # Details-Spalte umbricht
                        cell.alignment = align(
                            "center" if ci in (1, 3, 5, 6, 7, 8, 9, 11) else "left",
                            wrap=wrap)

                    # Gleitzeit einfärben
                    if ei == 0 and gleit != 0:
                        ws.cell(data_row, 10).font = Font(
                            name="Calibri", bold=True, size=10,
                            color=gleit_color)

                    # Mehrzeilige Tage: leichte Abgrenzung
                    if ei > 0:
                        ws.cell(data_row, 11).fill = fill("E3F2FD")

                    data_row += 1

            # ── Details-Spalte auto-breit (Spalte O = 15) ───────────────────
            max_detail_len = 30   # Mindestbreite
            for row_cells in ws.iter_rows(min_row=7, max_row=data_row - 1,
                                          min_col=15, max_col=15):
                for cell in row_cells:
                    if cell.value:
                        max_detail_len = max(max_detail_len, len(str(cell.value)) + 4)
            ws.column_dimensions["O"].width = min(max_detail_len, 80)  # max 80 Zeichen

            # Keine Summen-/Statistik-Zeile — Auswertung erfolgt in der App
            gleit_total_str = (f"+{gleit_h:.2f} h" if gleit_h > 0
                               else f"{gleit_h:.2f} h")

            # ── Seiteneinrichtung ─────────────────────────────────────────────
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize   = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage   = True
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows       = "6:6"   # Kopfzeile auf jeder Seite
            ws.oddHeader.center.text  = (
                f"Servicezeitenmeldung — {name} — {month_name} {year}")
            ws.oddFooter.right.text   = "Seite &P von &N"

            wb.save(dest)
            self.after(0, lambda: messagebox.showinfo(
                "Gespeichert",
                f"Servicezeitenmeldung gespeichert:\n{dest}\n\n"
                f"Außendienst: {ad_days} Tage\n"
                f"Innendienst: {id_days} Tage\n"
                f"Gesamt: {total_h:.2f} Stunden"))

        except Exception as e:
            err = sanitize_error(e)
            self.after(0, lambda: messagebox.showerror("Excel Fehler", err))

    # ── Excel: FB_0020 Reisekostenabrechnung ──────────────────────────────────

    def _generate_excel_reisekosten(self):
        tmpl = _get_template(TMPL_REISE, "FB_0020 Reisekostenabrechnung")
        if not tmpl:
            return
        kw, year = self._kw_var.get(), self._year_var.get()
        name_str = f"{self._v_nachname.get()}_{self._v_vorname.get()}"
        # Als .xlsx speichern → VBA-Makros werden entfernt, kein Laufzeitfehler 438
        dest = filedialog.asksaveasfilename(
            title="Reisekostenabrechnung speichern",
            initialfile=f"Reisekosten_KW{kw:02d}_{year}_{name_str}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Excel Makro", "*.xlsm"), ("Alle", "*.*")])
        if not dest:
            return
        # Nur aktuelle KW — aktuell geladene Daten + gespeicherte KW-Datei zusammenführen
        kw_file  = os.path.join(KW_DIR, f"{year}-W{kw:02d}.json")
        kw_data  = load_json(kw_file, {})
        kw_data.update(self._day_data)
        kw_data["_extras"] = {
            "sonstiges":    self._kw_sonstiges_var.get()     if hasattr(self, "_kw_sonstiges_var")     else "",
            "sonstiges_txt":self._kw_sonstiges_txt_var.get() if hasattr(self, "_kw_sonstiges_txt_var") else "",
        }
        threading.Thread(
            target=self._do_excel_reisekosten,
            args=(tmpl, dest, kw_data, kw, year), daemon=True).start()

    @staticmethod
    def _ws_set(ws, row, col, value, fmt=None):
        """Schreibt in eine Zelle — überspringt MergedCell (non-top-left) sicher."""
        try:
            from openpyxl.cell.cell import MergedCell
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                return   # Nur Top-Left einer Merge-Range ist beschreibbar
            cell.value = value
            if fmt:
                cell.number_format = fmt
        except Exception:
            pass

    def _do_excel_reisekosten(self, tmpl, dest, kw_data, kw, year):
        try:
            import openpyxl  # lazy – wird nur beim ersten Excel-Export geladen
            # Template laden (keep_vba=False → VBA wird entfernt, kein Fehler 438)
            shutil.copy(tmpl, dest)
            wb = openpyxl.load_workbook(dest, keep_vba=False)
            ws = wb.active

            name    = f"{self._v_nachname.get()}, {self._v_vorname.get()}".strip(", ")
            wohnort = self._v_wohnort.get() or "Wohnort"
            dates   = _week_dates(year, kw)

            # ── Header ────────────────────────────────────────────────────────
            self._ws_set(ws, 4, 2, name)          # B4
            self._ws_set(ws, 5, 2, "K&B Coding - Service")  # B5
            self._ws_set(ws, 7, 2, "Service")     # B7
            self._ws_set(ws, 7, 6, dates[0], "DD.MM.YY")   # F7 = KW-Start
            self._ws_set(ws, 7, 8, dates[-1], "DD.MM.YY")  # H7 = KW-Ende

            # ── KW-Extras (Sonstiges + km) aus kw_data extrahieren ───────────
            kw_extras = kw_data.get("_extras", {})   # .get statt .pop → kein Seiteneffekt

            # ── Datenblöcke: 3 Zeilen pro EINTRAG ab Zeile 12 ────────────────
            # Pro Eintrag (auch mehrere pro Tag):
            #   Zeile A: Datum, Startzeit, Übernachtung, Frühstück
            #   Zeile B: Endzeit, Details/Route, Land, Mittagessen
            #   Zeile C: Abendessen
            block = 0
            for ds in sorted(kw_data):
                day = kw_data[ds]
                if day.get("status", "Arbeit") != "Arbeit":
                    continue
                try:
                    d = datetime.date.fromisoformat(ds)
                except Exception:
                    continue
                entries = day.get("entries") or [{}]
                first_ad = True   # erstes Außendienst-Eintrag des Tages
                for ei, entry in enumerate(entries):
                    # Innendienst-Einträge nicht in Reisekosten
                    dienst = entry.get("dienst", "Außendienst")
                    if dienst in ("Innendienst", "Homeoffice", "Home Office"):
                        continue

                    ra = 12 + block * 3
                    rb = ra + 1
                    rc = ra + 2
                    if ra > 188:
                        break
                    # Per-Eintrag-Zeiten (Fallback auf Tag-Ebene für ersten AD-Eintrag)
                    e_start = entry.get("start") or (day.get("start","") if first_ad else "")
                    e_end   = entry.get("end")   or (day.get("end",  "") if first_ad else "")
                    land    = entry.get("land")  or day.get("land", "DE") or "DE"
                    reise   = _entry_detail(entry, wohnort)
                    t_s = _str_to_time(e_start)
                    t_e = _str_to_time(e_end)

                    # Zeile A: Datum (nur beim ersten AD-Eintrag des Tages)
                    if first_ad:
                        self._ws_set(ws, ra, 1, d, "DD.MM.YY")
                    if t_s:
                        self._ws_set(ws, ra, 2, t_s, "HH:MM")
                    # Reiseweg
                    if reise:
                        self._ws_set(ws, ra, 4, reise)
                    # Übernachtung/Früh nur beim ersten AD-Eintrag
                    if first_ad and day.get("uebernacht") == "ja":
                        self._ws_set(ws, ra, 9, "ja")
                    if first_ad and day.get("fruehstueck") == "ja":
                        self._ws_set(ws, ra, 10, "ja")

                    # Zeile B: Endzeit + Land
                    if t_e:
                        self._ws_set(ws, rb, 2, t_e, "HH:MM")
                    self._ws_set(ws, rb, 6, land)
                    if first_ad and day.get("mittag") == "ja":
                        self._ws_set(ws, rb, 10, "ja")

                    # Zeile C: Abend
                    if first_ad and day.get("abend") == "ja":
                        self._ws_set(ws, rc, 10, "ja")

                    first_ad = False
                    block += 1

            # ── KW-Sonstiges → G195 (Bezeichnung) + P195 (Betrag) ───────────
            try:
                s_txt = kw_extras.get("sonstiges_txt", "").strip()
                s_val = float(kw_extras.get("sonstiges", "") or 0)
                if s_val or s_txt:
                    self._ws_set(ws, 195, 7, s_txt or "Sonstige Kosten")
                    if s_val:
                        self._ws_set(ws, 195, 16, s_val)
            except Exception:
                pass

            wb.save(dest)
            self.after(0, lambda: messagebox.showinfo(
                "Gespeichert",
                f"Reisekostenabrechnung gespeichert:\n{dest}\n\n"
                f"KW {kw} / {year}  –  {block} Einträge"))
        except Exception as e:
            err = sanitize_error(e)
            self.after(0, lambda: messagebox.showerror("Excel Fehler", err))
